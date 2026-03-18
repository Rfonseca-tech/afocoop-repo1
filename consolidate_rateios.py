"""
consolidate_rateios.py
======================
Consolidates AFOOCOP monthly rateio files (DPA + FAP) into one master Excel workbook.

Output: AFOOCOP_Rateios_Consolidado.xlsx
  Sheet 1: MASTER_DATA   — all transactions, one row per line item
  Sheet 2: AUDIT_LOG     — source file/sheet processing summary
  Sheet 3: DATA_QUALITY  — data quality checks report
"""

import re
import sys
import warnings
from pathlib import Path
from datetime import datetime
from typing import Optional

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings("ignore", category=UserWarning)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
BASE = Path(r"C:\Users\Sairon\OneDrive - R FONSECA ADVOGADOS\01 - PROJETOS\AFOOCOP\Analise dos rateios")
OUTPUT_FILE = Path(r"C:\Users\Sairon\OneDrive - R FONSECA ADVOGADOS\01 - PROJETOS\AFOOCOP\AFOOCOP_Rateios_Consolidado.xlsx")

MONTH_MAP = {
    "janeiro": 1, "fevereiro": 2, "março": 3, "marco": 3,
    "abril": 4, "maio": 5, "junho": 6, "julho": 7,
    "agosto": 8, "setembro": 9, "outubro": 10,
    "novembro": 11, "dezembro": 12,
}

# Sheets to skip (summary/metadata, not transactional)
SKIP_SHEETS = {"resumo"}

# Sheet-type detection by name keywords
SHEET_TYPE_MAP = {
    "acidente": "Acidente",
    "complemento": "Complemento",
    "credito": "Credito",
    "crédito": "Credito",
    "assistencia": "Assistencia",
    "assistência": "Assistencia",
}

# Per sheet-type: candidate value column names (in priority order)
VALUE_COL_CANDIDATES = {
    "Acidente":    ["Total", "Valor"],
    "Complemento": ["Valor(comp)", "Valor(cred)", "Valor"],
    "Credito":     ["Valor(cred)", "Valor(comp)", "Valor"],
    "Assistencia": ["Valor(assist)", "Valor"],
}

# Per sheet-type: candidate ID column names
ID_COL_CANDIDATES = {
    "Acidente":    ["Acidente"],
    "Complemento": ["Complemento"],
    "Credito":     ["Crédito", "Credito", "CrÚdito"],
    "Assistencia": ["Assistencia", "Assistência"],
}

# Plate column names
PLACA_CANDIDATES = ["Placa", "placa", "PLACA"]

# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def infer_month(filename: str) -> Optional[datetime]:
    """Extract month/year from filename like 'Rateio - Janeiro - 2025 - DPA.xlsx'."""
    name_lower = filename.lower()
    for pt_month, month_num in MONTH_MAP.items():
        # Also handle accented chars that may have been mangled
        if pt_month in name_lower or pt_month.replace("ç", "c").replace("r", "r") in name_lower:
            year_match = re.search(r"(\d{4})", filename)
            if year_match:
                year = int(year_match.group(1))
                return datetime(year, month_num, 1)
    return None


def detect_col(df: pd.DataFrame, candidates: list[str]) -> Optional[str]:
    """Return first matching column name from candidates (case-insensitive)."""
    col_lower = {c.lower(): c for c in df.columns}
    for cand in candidates:
        if cand.lower() in col_lower:
            return col_lower[cand.lower()]
    return None


def normalize_placa(series: pd.Series) -> pd.Series:
    return series.astype(str).str.strip().str.upper().replace("NAN", pd.NA)


def detect_sheet_type(sheet_name: str) -> Optional[str]:
    """Map sheet name to a canonical type."""
    name_lower = sheet_name.lower().strip()
    for keyword, tipo in SHEET_TYPE_MAP.items():
        if keyword in name_lower:
            return tipo
    return None


def find_header_row(filepath: Path, sheet: str, max_scan: int = 10) -> int:
    """
    Find the header row index by looking for the first row that contains 'Placa'
    or another known column name. Falls back to row 1 (0-indexed).
    """
    try:
        df_raw = pd.read_excel(filepath, sheet_name=sheet, nrows=max_scan, header=None)
        for i, row in df_raw.iterrows():
            row_vals = [str(v).lower().strip() for v in row if pd.notna(v)]
            if "placa" in row_vals:
                return i
    except Exception:
        pass
    return 1  # default: second row (0-indexed)


# ---------------------------------------------------------------------------
# Core reader
# ---------------------------------------------------------------------------

def read_sheet(filepath: Path, sheet: str, fundo: str, month: datetime,
               month_label: str, audit_rows: list, quality_issues: list) -> Optional[pd.DataFrame]:
    """Read one transactional sheet and normalize it. Returns normalized DataFrame or None."""

    tipo = detect_sheet_type(sheet)
    if tipo is None:
        audit_rows.append({
            "SOURCE_FILE": filepath.name,
            "SOURCE_SHEET": sheet,
            "FUNDO": fundo,
            "MONTH_LABEL": month_label,
            "TIPO_LANCAMENTO": "DESCONHECIDO",
            "ROWS_READ": 0,
            "ROWS_EXCLUDED": 0,
            "STATUS": "SKIPPED — sheet type not recognized",
        })
        return None

    header_row = find_header_row(filepath, sheet)

    try:
        df = pd.read_excel(filepath, sheet_name=sheet, header=header_row)
    except Exception as e:
        audit_rows.append({
            "SOURCE_FILE": filepath.name,
            "SOURCE_SHEET": sheet,
            "FUNDO": fundo,
            "MONTH_LABEL": month_label,
            "TIPO_LANCAMENTO": tipo,
            "ROWS_READ": 0,
            "ROWS_EXCLUDED": 0,
            "STATUS": f"ERROR — {e}",
        })
        return None

    # Drop all-null rows and columns
    df = df.dropna(how="all").dropna(axis=1, how="all")
    rows_read = len(df)

    # ----- Plate column -----
    placa_col = detect_col(df, PLACA_CANDIDATES)
    if placa_col is None:
        audit_rows.append({
            "SOURCE_FILE": filepath.name,
            "SOURCE_SHEET": sheet,
            "FUNDO": fundo,
            "MONTH_LABEL": month_label,
            "TIPO_LANCAMENTO": tipo,
            "ROWS_READ": rows_read,
            "ROWS_EXCLUDED": rows_read,
            "STATUS": "SKIPPED — Placa column not found",
        })
        return None

    # ----- Value column -----
    value_col = detect_col(df, VALUE_COL_CANDIDATES[tipo])
    if value_col is None:
        audit_rows.append({
            "SOURCE_FILE": filepath.name,
            "SOURCE_SHEET": sheet,
            "FUNDO": fundo,
            "MONTH_LABEL": month_label,
            "TIPO_LANCAMENTO": tipo,
            "ROWS_READ": rows_read,
            "ROWS_EXCLUDED": rows_read,
            "STATUS": "SKIPPED — value column not found",
        })
        return None

    # ----- ID column -----
    id_col = detect_col(df, ID_COL_CANDIDATES[tipo])

    # ----- Associado column -----
    assoc_col = detect_col(df, ["Associado", "ASSOCIADO", "associado"])

    # ----- Data Rateio column -----
    rateio_col = detect_col(df, ["Data do Rateio", "Data de Rateio"])

    # ----- Data Evento column -----
    evento_col = detect_col(df, [
        "Data do Acidente", "Data da acidente", "Data do acidente",
        "Data(assist)", "Data da AssistÛncia", "Data da Assistência", "Data da AssistÙncia",
        "Data do assistÛncia", "Data do assistência",
        "Data do acidente original", "Data",
    ])

    # ----- Perda column -----
    perda_col = detect_col(df, ["Perda"])

    # ----- Build normalized DataFrame -----
    df["_PLACA_RAW"] = df[placa_col]
    df["PLACA"] = normalize_placa(df[placa_col])

    # Exclude rows where Placa is null / "TOTAL" / "NAN" / clearly not a plate
    mask_valid = (
        df["PLACA"].notna()
        & ~df["PLACA"].str.upper().isin(["TOTAL", "NAN", "NONE", "", "PLACA"])
        & (df["PLACA"].str.len() > 3)
    )
    rows_excluded = (~mask_valid).sum()
    df = df[mask_valid].copy()

    # Also exclude rows where the ID column value is 'TOTAL'
    if id_col:
        id_total_mask = df[id_col].astype(str).str.upper().str.strip() == "TOTAL"
        rows_excluded += id_total_mask.sum()
        df = df[~id_total_mask].copy()

    # Coerce value to numeric
    df["VALOR"] = pd.to_numeric(df[value_col], errors="coerce")

    # Creditos: mark as negative if value is positive (they are outflows)
    # NOTE: some files already store them as negative; preserve sign if so.
    # If all values positive, negate them.
    if tipo == "Credito":
        if df["VALOR"].notna().any() and (df["VALOR"] > 0).all():
            df["VALOR"] = -df["VALOR"]

    out = pd.DataFrame(index=df.index)
    out["MONTH"] = month
    out["MONTH_LABEL"] = month_label
    out["FUNDO"] = fundo
    out["PLACA"] = df["PLACA"]
    out["ASSOCIADO"] = df[assoc_col].astype(str).str.strip() if assoc_col else pd.NA
    out["TIPO_LANCAMENTO"] = tipo
    out["ID_LANCAMENTO"] = df[id_col].astype(str).str.strip() if id_col else pd.NA
    out["VALOR"] = df["VALOR"]
    out["DATA_RATEIO"] = pd.to_datetime(df[rateio_col], errors="coerce") if rateio_col else pd.NaT
    out["DATA_EVENTO"] = pd.to_datetime(df[evento_col], errors="coerce") if evento_col else pd.NaT
    out["PERDA"] = df[perda_col].astype(str).str.strip() if perda_col else pd.NA
    out["SOURCE_FILE"] = filepath.name
    out["SOURCE_SHEET"] = sheet
    out["NOTES"] = ""

    # Flag zero/null values
    null_valor = out["VALOR"].isna()
    zero_valor = out["VALOR"] == 0
    if null_valor.any():
        quality_issues.append({
            "CHECK": "VALOR_NULO",
            "FILE": filepath.name,
            "SHEET": sheet,
            "COUNT": int(null_valor.sum()),
            "DETAIL": f"Linhas com VALOR nulo: {null_valor.sum()}",
        })
    if zero_valor.any():
        quality_issues.append({
            "CHECK": "VALOR_ZERO",
            "FILE": filepath.name,
            "SHEET": sheet,
            "COUNT": int(zero_valor.sum()),
            "DETAIL": f"Linhas com VALOR = 0: {zero_valor.sum()}",
        })

    audit_rows.append({
        "SOURCE_FILE": filepath.name,
        "SOURCE_SHEET": sheet,
        "FUNDO": fundo,
        "MONTH_LABEL": month_label,
        "TIPO_LANCAMENTO": tipo,
        "ROWS_READ": rows_read,
        "ROWS_EXCLUDED": rows_excluded,
        "STATUS": f"OK — {len(out)} linhas incluídas",
    })

    return out if len(out) > 0 else None


# ---------------------------------------------------------------------------
# Main consolidation
# ---------------------------------------------------------------------------

def consolidate():
    all_frames = []
    audit_rows = []
    quality_issues = []

    for fundo_folder in ["DPA", "FAP"]:
        folder = BASE / fundo_folder
        if not folder.exists():
            print(f"[SKIP] Pasta não encontrada: {folder}")
            continue

        files = sorted(folder.glob("*.xlsx")) + sorted(folder.glob("*.xls"))
        print(f"\n=== FUNDO: {fundo_folder} — {len(files)} arquivo(s) ===")

        for filepath in files:
            month = infer_month(filepath.name)
            if month is None:
                print(f"  [WARN] Não foi possível inferir mês: {filepath.name}")
                audit_rows.append({
                    "SOURCE_FILE": filepath.name,
                    "SOURCE_SHEET": "—",
                    "FUNDO": fundo_folder,
                    "MONTH_LABEL": "DESCONHECIDO",
                    "TIPO_LANCAMENTO": "—",
                    "ROWS_READ": 0,
                    "ROWS_EXCLUDED": 0,
                    "STATUS": "SKIPPED — mês não identificado no nome do arquivo",
                })
                continue

            month_label = f"{month.strftime('%B').capitalize()}/{month.year}"
            # Use Portuguese month name
            pt_months = {1:'Janeiro',2:'Fevereiro',3:'Março',4:'Abril',5:'Maio',6:'Junho',
                         7:'Julho',8:'Agosto',9:'Setembro',10:'Outubro',11:'Novembro',12:'Dezembro'}
            month_label = f"{pt_months[month.month]}/{month.year}"

            # Read sheet names
            try:
                wb_probe = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
                sheets = wb_probe.sheetnames
                wb_probe.close()
            except Exception as e:
                print(f"  [ERROR] {filepath.name}: {e}")
                audit_rows.append({
                    "SOURCE_FILE": filepath.name,
                    "SOURCE_SHEET": "—",
                    "FUNDO": fundo_folder,
                    "MONTH_LABEL": month_label,
                    "TIPO_LANCAMENTO": "—",
                    "ROWS_READ": 0,
                    "ROWS_EXCLUDED": 0,
                    "STATUS": f"ERROR ao abrir — {e}",
                })
                continue

            print(f"  {filepath.name}  [{month_label}]  sheets: {sheets}")

            for sheet in sheets:
                if sheet.lower().strip() in SKIP_SHEETS:
                    audit_rows.append({
                        "SOURCE_FILE": filepath.name,
                        "SOURCE_SHEET": sheet,
                        "FUNDO": fundo_folder,
                        "MONTH_LABEL": month_label,
                        "TIPO_LANCAMENTO": "Resumo",
                        "ROWS_READ": 0,
                        "ROWS_EXCLUDED": 0,
                        "STATUS": "SKIPPED — sheet Resumo (sumário, não transacional)",
                    })
                    continue

                df_sheet = read_sheet(filepath, sheet, fundo_folder, month,
                                      month_label, audit_rows, quality_issues)
                if df_sheet is not None:
                    all_frames.append(df_sheet)

    if not all_frames:
        print("\n[ERRO] Nenhum dado foi consolidado.")
        sys.exit(1)

    master = pd.concat(all_frames, ignore_index=True)
    master = master.sort_values(["MONTH", "FUNDO", "TIPO_LANCAMENTO", "PLACA"]).reset_index(drop=True)

    # ---------------------------------------------------------------------------
    # Merge with Unified Vehicle Database
    # ---------------------------------------------------------------------------
    veiculos_file = BASE.parent / "AFOOCOP_Base_Veiculos_Unificada.xlsx"
    if veiculos_file.exists():
        print(f"\n=== CRUZAMENTO: Lendo {veiculos_file.name} ===")
        try:
            df_veic = pd.read_excel(veiculos_file)
            
            # Normalize Placa to match master
            if "PLACA" in df_veic.columns:
                df_veic["PLACA"] = normalize_placa(df_veic["PLACA"])
                
                # First, drop completely empty/invalid plates from the base
                valid_plate = df_veic["PLACA"].notna() & (df_veic["PLACA"].str.len() > 3)
                df_veic = df_veic[valid_plate].copy()
                
                df_veic_dedup = df_veic.drop_duplicates(subset=["PLACA"], keep="last").copy()
                
                # Log number of total unique plates found in the base
                print(f"  Encontramos {len(df_veic_dedup)} placas únicas cadastradas em {veiculos_file.name}.")
                
                # Select columns to bring over
                cols_to_bring = ["PLACA"]
                potential_cols = ["Cavalo/Carreta", "Marca", "Modelo", "Ano Fabricação", "Ano Modelo", "VALOR_EQUIPAMENTO", "Status Processo", "Associado"]
                
                # Rename dict to standardize column names in master
                rename_dict = {
                    "VALOR_EQUIPAMENTO": "Valor Equipamento"
                }
                
                for c in potential_cols:
                    if c in df_veic_dedup.columns:
                        cols_to_bring.append(c)
                        if c == "Associado":
                            rename_dict["Associado"] = "ASSOCIADO_BASE"
                
                df_veic_dedup = df_veic_dedup[cols_to_bring].rename(columns=rename_dict)
                
                # Merge (Left Join)
                linhas_antes = len(master)
                master = master.merge(df_veic_dedup, on="PLACA", how="left")
                
                # Reposition notes to the end
                if "NOTES" in master.columns:
                    notes_col = master.pop("NOTES")
                    master["NOTES"] = notes_col

                print(f"  Merge concluído de {linhas_antes} linhas do rateio consolidado.")
                
            else:
                print("  [WARN] Coluna 'Placa' não encontrada na base de veículos.")
        except Exception as e:
            print(f"  [ERROR] Falha ao ler base de veículos: {e}")
    else:
        print(f"\n[WARN] Base de veículos não encontrada: {veiculos_file}")

    # ---------------------------------------------------------------------------
    # Data quality checks
    # ---------------------------------------------------------------------------
    print(f"\n=== MASTER_DATA: {len(master):,} linhas totais ===")

    # Check 1: duplicate PLACA + MONTH + FUNDO + TIPO (expected and fine — multiple events per plate)
    dupes = master.groupby(["PLACA", "MONTH", "FUNDO"]).size().reset_index(name="count")
    dupes_multi = dupes[dupes["count"] > 1]
    quality_issues.insert(0, {
        "CHECK": "PLACAS_MULTIPLOS_LANCAMENTOS",
        "FILE": "MASTER",
        "SHEET": "—",
        "COUNT": int(len(dupes_multi)),
        "DETAIL": f"Placas com >1 lançamento no mesmo mês/fundo (esperado): {len(dupes_multi)}",
    })

    # Check 2: missing PLACA
    null_placa = master["PLACA"].isna().sum()
    quality_issues.insert(1, {
        "CHECK": "PLACA_NULA",
        "FILE": "MASTER",
        "SHEET": "—",
        "COUNT": int(null_placa),
        "DETAIL": f"Linhas com PLACA nula: {null_placa}",
    })

    # Check 3: missing MONTH
    null_month = master["MONTH"].isna().sum()
    quality_issues.insert(2, {
        "CHECK": "MONTH_NULO",
        "FILE": "MASTER",
        "SHEET": "—",
        "COUNT": int(null_month),
        "DETAIL": f"Linhas com MONTH nulo: {null_month}",
    })

    # Check 4: months covered
    months_covered = sorted(master["MONTH"].dropna().unique())
    quality_issues.insert(3, {
        "CHECK": "MESES_COBERTOS",
        "FILE": "MASTER",
        "SHEET": "—",
        "COUNT": len(months_covered),
        "DETAIL": ", ".join([pd.Timestamp(m).strftime("%b/%Y") for m in months_covered]),
    })

    # Check 5: totals by fundo/tipo
    summary = master.groupby(["FUNDO", "TIPO_LANCAMENTO"])["VALOR"].agg(
        COUNT="count", TOTAL="sum"
    ).reset_index()
    for _, row in summary.iterrows():
        quality_issues.append({
            "CHECK": "TOTAL_POR_FUNDO_TIPO",
            "FILE": "MASTER",
            "SHEET": "—",
            "COUNT": int(row["COUNT"]),
            "DETAIL": f"FUNDO={row['FUNDO']}, TIPO={row['TIPO_LANCAMENTO']}, TOTAL_VALOR=R${row['TOTAL']:,.2f}",
        })

    # ---------------------------------------------------------------------------
    # Build DataFrames for output sheets
    # ---------------------------------------------------------------------------
    audit_df = pd.DataFrame(audit_rows)
    quality_df = pd.DataFrame(quality_issues)

    # ---------------------------------------------------------------------------
    # Write to Excel with formatting
    # ---------------------------------------------------------------------------
    print(f"\n  Salvando em: {OUTPUT_FILE}")

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        master.to_excel(writer, sheet_name="MASTER_DATA", index=False)
        audit_df.to_excel(writer, sheet_name="AUDIT_LOG", index=False)
        quality_df.to_excel(writer, sheet_name="DATA_QUALITY", index=False)

    # Apply formatting
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    _format_workbook(wb, master, audit_df, quality_df)
    wb.save(OUTPUT_FILE)

    print(f"\n[OK] Arquivo gerado com sucesso: {OUTPUT_FILE.name}")
    print(f"   MASTER_DATA : {len(master):,} linhas")
    print(f"   AUDIT_LOG   : {len(audit_df):,} entradas")
    print(f"   DATA_QUALITY: {len(quality_df):,} checks")


def _format_workbook(wb, master_df, audit_df, quality_df):
    """Apply basic formatting: freeze panes, column widths, header styling."""

    HEADER_FILL = PatternFill("solid", fgColor="1F3864")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
    CREDITO_FILL = PatternFill("solid", fgColor="FCE4D6")
    ALT_FILL = PatternFill("solid", fgColor="EEF2F8")
    THIN = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.freeze_panes = "A2"

        # Style header row
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER

        # Auto column widths (capped at 50)
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) if cell.value is not None else 0)
                for cell in col
            )
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        # Alternating row fill + highlight Credito rows (MASTER_DATA only)
        if sheet_name == "MASTER_DATA":
            tipo_col_idx = None
            valor_col_idx = None
            for i, cell in enumerate(ws[1], 1):
                if str(cell.value) == "TIPO_LANCAMENTO":
                    tipo_col_idx = i
                if str(cell.value) == "VALOR":
                    valor_col_idx = i

            for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
                tipo_val = row[tipo_col_idx - 1].value if tipo_col_idx else None
                is_credito = tipo_val == "Credito"
                for cell in row:
                    cell.border = BORDER
                    if is_credito:
                        cell.fill = CREDITO_FILL
                    elif row_idx % 2 == 0:
                        cell.fill = ALT_FILL
                if valor_col_idx:
                    valor_cell = row[valor_col_idx - 1]
                    valor_cell.number_format = "#,##0.00"

        # Format MONTH column as date
        if sheet_name == "MASTER_DATA":
            month_col_idx = None
            for i, cell in enumerate(ws[1], 1):
                if str(cell.value) == "MONTH":
                    month_col_idx = i
                    break
            if month_col_idx:
                for row in ws.iter_rows(min_row=2, min_col=month_col_idx, max_col=month_col_idx):
                    for cell in row:
                        cell.number_format = "MMM/YYYY"

        # Set row height for header
        ws.row_dimensions[1].height = 30

    # Tab colors
    if "MASTER_DATA" in wb.sheetnames:
        wb["MASTER_DATA"].sheet_properties.tabColor = "1F3864"
    if "AUDIT_LOG" in wb.sheetnames:
        wb["AUDIT_LOG"].sheet_properties.tabColor = "70AD47"
    if "DATA_QUALITY" in wb.sheetnames:
        wb["DATA_QUALITY"].sheet_properties.tabColor = "ED7D31"


# ---------------------------------------------------------------------------
if __name__ == "__main__":
    consolidate()
